from imports import *

class DataSplitter:
    def __init__(self, part):
        self.part = part

    def open_dir(self, folder, train, test):
        images_folder = os.path.join("", folder)
        if not os.path.exists(images_folder):
            os.makedirs(images_folder)
            print(images_folder)

        animal_images = os.path.join(folder, "animal")
        if not os.path.exists(animal_images):
            os.makedirs(animal_images)
            print(animal_images)

        not_animal_images = os.path.join(folder, "not_animal")
        if not os.path.exists(not_animal_images):
            os.makedirs(not_animal_images)
            print(not_animal_images)

        train_folder = os.path.join("", train)
        if not os.path.exists(train_folder):
            os.makedirs(train_folder)
            print(train_folder)

        test_folder = os.path.join("", test)
        if not os.path.exists(test_folder):
            os.makedirs(test_folder)
            print(test_folder)

        for folder_name in ["animal", "not_animal"]:
            new_path_train = os.path.join(train, folder_name)
            if not os.path.exists(new_path_train):
                os.makedirs(new_path_train)
                print(new_path_train)

        for folder_name in ["animal", "not_animal"]:
            new_path_test = os.path.join(test, folder_name)
            if not os.path.exists(new_path_test):
                os.makedirs(new_path_test)
                print(new_path_test)


    def move_images(self, excel_name, src_folder ,animal_folder, not_animal_folder):
        # import module
        import openpyxl
        # load excel with its path
        wrkbk = openpyxl.load_workbook(excel_name)
        sh = wrkbk.active
        x = 0 #0 for not animal, 1 for animal
        animal_photo_base_list = []
        not_animal_photo_base_list = []
        # iterate through excel and display data
        for i in range(1, sh.max_row + 1):
            # print("\n")
            # print("Row ", i, " data :")
            for j in range(1, sh.max_column + 1):
                cell_obj = sh.cell(row=i, column=j)
                if j==4:
                    if cell_obj.value == "Lion" or cell_obj.value == "Sphinx":
                        # print(cell_obj.value, end=" ")
                        x=1
                    else:
                        # print(cell_obj.value, end=" ")
                        x=0
                if j==5:
                    if x == 1:
                        files_names = str(cell_obj.value).split("tif\n")
                        files_names[:] = [x + "tif" for x in files_names]
                        files_names.pop()
                        photo_base = [x for x in files_names if "photoBase" in x]
                        animal_photo_base_list.append(photo_base)
                        # print(files_names)
                        # print("Animal")
                    else:
                        files_names = str(cell_obj.value).split("tif\n")
                        files_names[:] = [x + "tif" for x in files_names]
                        files_names.pop()
                        photo_base = [x for x in files_names if "photoBase" in x]
                        not_animal_photo_base_list.append(photo_base)

                        # files_names = cell_obj.value
                        # sh.cell(row=i, column=j).value = files_names + " not_animal"
                        # print(files_names)
                        # print("not animal")
        # print(animal_photo_base_list)
        images = os.listdir(src_folder)
        for f in images:
            file_name = str(f)
            # print(file_name)
            src = os.path.join(src_folder, f)

            if any(file_name in s for s in animal_photo_base_list):
                dst = os.path.join(animal_folder, f)
                move(src, animal_folder)
            if any(file_name in s for s in not_animal_photo_base_list):
                dst = os.path.join(not_animal_folder, f)
                move(src, not_animal_folder)
                # print("src is " + src)
                # print("dst is " + dst)

    def split(self, folder, train, test):
        n = 0
        for i in ["animal", "not_animal"]:
            folder_name = i
            current_dir = os.path.join(folder, folder_name)
            print(current_dir)
            j = 0
            for image_path in os.listdir(current_dir):
                src = os.path.join(current_dir, image_path)
                print(src)
                if (j % self.part) == 0:
                    # if(n % 1) == 0:
                    dst_dir = os.path.join(test, folder_name)
                    dst = os.path.join(dst_dir, image_path)
                    print(dst)
                    print(j)
                    n += 1
                    # else:
                    #     dst_dir = os.path.join(val, folder_name)
                    #     dst = os.path.join(dst_dir, image_path)
                    #     print(dst)
                    #     print(j)
                    #     n+=1
                else:
                    dst_dir = os.path.join(train, folder_name)
                    dst = os.path.join(dst_dir, image_path)
                    print(dst)
                    print(j)
                try:
                    shutil.copy(src, dst)
                except EnvironmentError:
                    print(n)
                j += 1
                if (j == 11):
                    j = 0