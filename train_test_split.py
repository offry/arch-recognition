from imports import *

import shutil
import os
import pickle


def find_diff_drawing_and_photo(drawing_dir, photo_dir):
    drawing_list, photo_list = [], []
    drawing_dir = os.path.join(os.getcwd(), drawing_dir)
    for label in os.listdir(drawing_dir):
        if label=="Sphinx" or label=="Duck":
            continue
        current_dir = os.path.join(drawing_dir, label)
        drawing_list.extend(os.listdir(current_dir))
    photo_dir = os.path.join(os.getcwd(), photo_dir)
    for label in os.listdir(photo_dir):
        if label=="Sphinx" or label=="Duck":
            continue
        current_dir = os.path.join(photo_dir, label)
        photo_list.extend(os.listdir(current_dir))
    drawing_list = [i.split('_')[0] for i in drawing_list]
    photo_list = [i.split('_')[0] for i in photo_list]
    not_in_photo = list(set(drawing_list)-set(photo_list))
    not_in_drawing = list(set(photo_list)-set(drawing_list))

    for i in not_in_photo:
        print("not in photo: {}".format(i))
    for i in not_in_drawing:
        print("not in drawing: {}".format(i))


def create_all_data_drawing(all_data_drawing, drawing_dir):
    drawing_dir = os.path.join(os.getcwd(), drawing_dir)
    all_data_drawing = os.path.join(os.getcwd(), all_data_drawing)
    if not os.path.isdir(all_data_drawing):
        os.mkdir(all_data_drawing)
    for label in os.listdir(drawing_dir):
        if label=="Sphinx" or label=="Duck":
            continue
        current_dir = os.path.join(drawing_dir, label)
        for file in os.listdir(current_dir):
            src = os.path.join(current_dir, file)
            dst = os.path.join(all_data_drawing, file)
            print(file)
            shutil.copy2(src, dst)


def create_all_data_photo(all_data_photo, photo_dir):
    photo_dir = os.path.join(os.getcwd(), photo_dir)
    all_data_photo = os.path.join(os.getcwd(), all_data_photo)
    if not os.path.isdir(all_data_photo):
        os.mkdir(all_data_photo)
    for label in os.listdir(photo_dir):
        if label=="Sphinx" or label=="Duck":
            continue
        current_dir = os.path.join(photo_dir, label)
        for file in os.listdir(current_dir):
            src = os.path.join(current_dir, file)
            dst = os.path.join(all_data_photo, file)
            print(file)
            shutil.copy2(src, dst)


def update_labels_excel(excel_name, images_dir):
    import openpyxl
    label_wrkbk = openpyxl.load_workbook(excel_name)
    label_sh = label_wrkbk.active
    labels_dir_list = os.listdir(images_dir)
    i = 2
    for label_dir in labels_dir_list:
        current_dir = os.path.join(images_dir, label_dir)
        for file in os.listdir(current_dir):
            image_name_obj = label_sh.cell(row=i, column=1)
            image_name_obj.value = file
            image_label_obj = label_sh.cell(row=i, column=2)
            image_label_obj.value = label_dir
            i+=1
    label_wrkbk.save(excel_name)


def update_photo_dirs(drawing_dir, photo_dir):
    drawing_dir_list = os.listdir(drawing_dir)
    photo_dir_list = os.listdir(photo_dir)
    new_all_data_dir = os.path.join(os.getcwd(),os.path.join("images_final", "drawing_base_all"))
    for dir in drawing_dir_list:
        dir_list = os.listdir(os.path.join(os.getcwd(), os.path.join(drawing_dir, dir)))
        for current_dir in dir_list:
            photo_current_dir = os.path.join(photo_dir, os.path.join(dir, current_dir))
            if not os.path.isdir(photo_current_dir):
                os.mkdir(photo_current_dir)
            for file in os.listdir(os.path.join(drawing_dir, os.path.join(dir, current_dir))):
                if "/" in file:
                    image_name_and_number = file.split("/")[-1].split("_")[0]
                else:
                    image_name_and_number = file.split("_")[0]
                for photo_file in os.listdir(new_all_data_dir):
                    if image_name_and_number in photo_file:
                        if "draw" not in photo_file:
                            break
                src = os.path.join(new_all_data_dir, photo_file)
                dst = os.path.join(photo_current_dir, photo_file)
                shutil.copy2(src, dst)
                print("copy {} to photo {}".format(image_name_and_number, dir + "_" + current_dir))

def update_photo_as_drawing_dir(drawing_dir, photo_dir):
    drawing_dir_list = os.listdir(drawing_dir)
    photo_dir_list = os.listdir(photo_dir)
    new_all_data_dir = os.path.join(os.getcwd(),os.path.join("images_final", "all_photo_as_drawing"))
    for dir in drawing_dir_list:
        dir_list = os.listdir(os.path.join(os.getcwd(), os.path.join(drawing_dir, dir)))
        for current_dir in dir_list:
            photo_current_dir = os.path.join(photo_dir, os.path.join(dir, current_dir))
            if not os.path.isdir(photo_current_dir):
                os.mkdir(photo_current_dir)
            for file in os.listdir(os.path.join(drawing_dir, os.path.join(dir, current_dir))):
                if "/" in file:
                    image_name_and_number = file.split("/")[-1].split("_")[0]
                else:
                    image_name_and_number = file.split("_")[0]
                for photo_file in os.listdir(new_all_data_dir):
                    if image_name_and_number in photo_file:
                        if "draw" not in photo_file:
                            break
                src = os.path.join(new_all_data_dir, photo_file)
                dst = os.path.join(photo_current_dir, photo_file)
                shutil.copy2(src, dst)
                print("copy {} to photo {}".format(image_name_and_number, dir + "_" + current_dir))



def create_labels_dict(excel_name, labels_dict_name):
    import openpyxl
    label_wrkbk = openpyxl.load_workbook(excel_name)
    label_sh = label_wrkbk.active
    filename = labels_dict_name
    print("Creating labels_dict.bin")
    with open(filename, 'wb') as outfile:
        labels_dict = {}
        for i in range(2, label_sh.max_row + 1):
            image_name_obj = label_sh.cell(row=i, column=1)
            image_label_obj = label_sh.cell(row=i, column=2)
            label = image_label_obj.value
            image_name = str(image_name_obj.value)
            labels_dict[image_name] = str(label)
        pickle.dump(labels_dict, outfile)


def check_image_label(given_image_path, labels_dict_name):
    filename = labels_dict_name
    with open(filename, 'rb') as infile:
        labels_dict = pickle.load(infile)
    if "/" in given_image_path:
        image_name_and_number = given_image_path.split("/")[-1].split("_")[0]
    else:
        image_name_and_number = given_image_path.split("_")[0]
    # iterate through excel and display data
    for image in labels_dict.keys():
        if image_name_and_number in image:
            return str(labels_dict[image])


def create_drawing_base_dirs(input_dir, photo_dir):
    photo_dir_list = os.listdir(photo_dir)
    input_dir_list = os.listdir(input_dir)

    labels_drawing_dir = os.path.join(os.path.join(os.getcwd(), "images_final"), "labeled_drawing")
    if not os.path.isdir(labels_drawing_dir):
        os.mkdir(labels_drawing_dir)

    for file in photo_dir_list:
        file_split_name = file.split("photoBase")
        drawing_file_name = [x for x in input_dir_list if file_split_name[0] in x and "draw" in x]
        if len(drawing_file_name)==1:
            drawing_label = check_image_label(file)
            if drawing_label is not None:
                label_dir = os.path.join(labels_drawing_dir, drawing_label)
                if not os.path.isdir(label_dir):
                    os.mkdir(label_dir)
                src_file = os.path.join(input_dir, drawing_file_name[0])
                dst_file = os.path.join(label_dir, drawing_file_name[0])
                if not os.path.isfile(dst_file):
                    if not os.path.isfile(os.path.join("images_final/drawing_not_use", drawing_label)):
                        shutil.copy2(src_file, dst_file)
                        print("copied {}".format(drawing_file_name[0]))


def calculate_how_many_images_in_each_label(labels_dict_name):
    filename = labels_dict_name
    with open(filename, 'rb') as infile:
        labels_dict = pickle.load(infile)
    labels_count_dict = {}
    for image in labels_dict.keys():
        if str(labels_dict[image]) in labels_count_dict:
            labels_count_dict[str(labels_dict[image])] = labels_count_dict[str(labels_dict[image])] + 1
        else:
            labels_count_dict[str(labels_dict[image])] = 1
    return labels_count_dict


def create_test_train_experiment(experiment_name, draw_dir, photo_dir, photo_as_draw_dir, supervised, test_size):
    if not os.path.isdir(os.path.join(photo_dir, "experiments")):
        os.mkdir(os.path.join(photo_dir, "experiments"))
    if not os.path.isdir(os.path.join(draw_dir, "experiments")):
        os.mkdir(os.path.join(draw_dir, "experiments"))
    if not os.path.isdir(os.path.join(photo_as_draw_dir, "experiments")):
        os.mkdir(os.path.join(photo_as_draw_dir, "experiments"))

    exp_photo_dir = os.path.join(photo_dir, os.path.join("experiments", experiment_name))
    if not os.path.isdir(exp_photo_dir):
        os.mkdir(exp_photo_dir)
    exp_draw_dir = os.path.join(draw_dir, os.path.join("experiments", experiment_name))
    if not os.path.isdir(exp_draw_dir):
        os.mkdir(exp_draw_dir)
    exp_photo_as_draw_dir = os.path.join(photo_as_draw_dir, os.path.join("experiments", experiment_name))
    if not os.path.isdir(exp_photo_as_draw_dir):
        os.mkdir(exp_photo_as_draw_dir)

    train_photo_dir = os.path.join(exp_photo_dir, "train")
    test_photo_dir = os.path.join(exp_photo_dir, "test")
    os.mkdir(train_photo_dir)
    os.mkdir(os.path.join(train_photo_dir, "1"))
    os.mkdir(test_photo_dir)
    train_draw_dir = os.path.join(exp_draw_dir, "train")
    test_draw_dir = os.path.join(exp_draw_dir, "test")
    os.mkdir(train_draw_dir)
    os.mkdir(os.path.join(train_draw_dir, "1"))
    os.mkdir(test_draw_dir)
    # train_photo_as_draw_dir = os.path.join(exp_photo_as_draw_dir, "train")
    # test_photo_as_draw_dir = os.path.join(exp_photo_as_draw_dir, "test")
    # os.mkdir(train_photo_as_draw_dir)
    # os.mkdir(os.path.join(train_photo_as_draw_dir, "1"))
    # os.mkdir(test_photo_as_draw_dir)

    if supervised:
        if not os.path.isdir(os.path.join(photo_dir, "supervised_experiments")):
            os.mkdir(os.path.join(photo_dir, "supervised_experiments"))
        supervised_exp_photo_dir = os.path.join(photo_dir, os.path.join("supervised_experiments", experiment_name))
        if not os.path.isdir(supervised_exp_photo_dir):
            os.mkdir(supervised_exp_photo_dir)
        supervised_train_photo_dir = os.path.join(supervised_exp_photo_dir, "train")
        supervised_test_photo_dir = os.path.join(supervised_exp_photo_dir, "test")
        os.mkdir(supervised_train_photo_dir)
        os.mkdir(supervised_test_photo_dir)

    if supervised:
        if not os.path.isdir(os.path.join(draw_dir, "supervised_experiments")):
            os.mkdir(os.path.join(draw_dir, "supervised_experiments"))
        supervised_exp_draw_dir = os.path.join(draw_dir, os.path.join("supervised_experiments", experiment_name))
        if not os.path.isdir(supervised_exp_draw_dir):
            os.mkdir(supervised_exp_draw_dir)
        supervised_train_draw_dir = os.path.join(supervised_exp_draw_dir, "train")
        supervised_test_draw_dir = os.path.join(supervised_exp_draw_dir, "test")
        os.mkdir(supervised_train_draw_dir)
        os.mkdir(supervised_test_draw_dir)

    # if supervised:
    #     if not os.path.isdir(os.path.join(photo_as_draw_dir, "supervised_experiments")):
    #         os.mkdir(os.path.join(photo_as_draw_dir, "supervised_experiments"))
    #     supervised_exp_photo_as_draw_dir = os.path.join(photo_as_draw_dir, os.path.join("supervised_experiments", experiment_name))
    #     if not os.path.isdir(supervised_exp_photo_as_draw_dir):
    #         os.mkdir(supervised_exp_photo_as_draw_dir)
    #     supervised_train_photo_as_draw_dir = os.path.join(supervised_exp_photo_as_draw_dir, "train")
    #     supervised_test_photo_as_draw_dir = os.path.join(supervised_exp_photo_as_draw_dir, "test")
    #     os.mkdir(supervised_train_photo_as_draw_dir)
    #     os.mkdir(supervised_test_photo_as_draw_dir)


    labels_count_dict = calculate_how_many_images_in_each_label("labels_dict_photo.bin")
    images_photo_dir = os.path.join(photo_dir, "labels")
    images_draw_dir = os.path.join(draw_dir, "labels")
    # images_photo_as_draw_dir = os.path.join(photo_as_draw_dir, "labels")
    labels_dir_list = os.listdir(images_photo_dir)
    for label_dir in labels_dir_list:
        if label_dir=="sphinx" or label_dir=="duck":
            continue
        current_photo_dir = os.path.join(images_photo_dir, label_dir)
        current_draw_dir = os.path.join(images_draw_dir, label_dir)
        # current_photo_as_draw_dir = os.path.join(images_photo_as_draw_dir, label_dir)

        label_number_of_images_for_test = round(labels_count_dict[str(label_dir)]*test_size)
        label_images = random.sample(range(0, labels_count_dict[str(label_dir)]), label_number_of_images_for_test)

        current_photo_list_dir = sorted(os.listdir(current_photo_dir))
        current_draw_list_dir = sorted(os.listdir(current_draw_dir))
        # current_photo_as_draw_list_dir = sorted(os.listdir(current_photo_as_draw_dir))

        for i, file in zip(range(len(current_photo_list_dir)), current_photo_list_dir):
            src_file_path = os.path.join(current_photo_dir, file)
            if i in label_images:
                dest_file_path = os.path.join(test_photo_dir, file)
                shutil.copy2(src_file_path, dest_file_path)
                if supervised:
                    label_supervised_test_dir = os.path.join(supervised_test_photo_dir, label_dir)
                    if not os.path.isdir(label_supervised_test_dir):
                        os.mkdir(label_supervised_test_dir)
                    supervised_dest_file_path = os.path.join(label_supervised_test_dir, file)
                    shutil.copy2(src_file_path, supervised_dest_file_path)
                print("{} number {} to test".format(str(label_dir), i))
            else:
                dest_file_path = os.path.join(os.path.join(train_photo_dir, "1"), file)
                shutil.copy2(src_file_path, dest_file_path)
                if supervised:
                    label_supervised_train_dir = os.path.join(supervised_train_photo_dir, label_dir)
                    if not os.path.isdir(label_supervised_train_dir):
                        os.mkdir(label_supervised_train_dir)
                    supervised_dest_file_path = os.path.join(label_supervised_train_dir, file)
                    shutil.copy2(src_file_path, supervised_dest_file_path)
                print("{} number {} to train".format(str(label_dir), i))
        for i, file in zip(range(len(current_draw_list_dir)), current_draw_list_dir):
            src_file_path = os.path.join(current_draw_dir, file)
            if i in label_images:
                dest_file_path = os.path.join(test_draw_dir, file)
                shutil.copy2(src_file_path, dest_file_path)
                if supervised:
                    label_supervised_test_dir = os.path.join(supervised_test_draw_dir, label_dir)
                    if not os.path.isdir(label_supervised_test_dir):
                        os.mkdir(label_supervised_test_dir)
                    supervised_dest_file_path = os.path.join(label_supervised_test_dir, file)
                    shutil.copy2(src_file_path, supervised_dest_file_path)
                print("{} number {} to test".format(str(label_dir), i))
            else:
                dest_file_path = os.path.join(os.path.join(train_draw_dir, "1"), file)
                shutil.copy2(src_file_path, dest_file_path)
                if supervised:
                    label_supervised_train_dir = os.path.join(supervised_train_draw_dir, label_dir)
                    if not os.path.isdir(label_supervised_train_dir):
                        os.mkdir(label_supervised_train_dir)
                    supervised_dest_file_path = os.path.join(label_supervised_train_dir, file)
                    shutil.copy2(src_file_path, supervised_dest_file_path)
                print("{} number {} to train".format(str(label_dir), i))
        # for i, file in zip(range(len(current_photo_as_draw_list_dir)), current_photo_as_draw_list_dir):
        #     src_file_path = os.path.join(current_photo_as_draw_dir, file)
        #     if i in label_images:
        #         dest_file_path = os.path.join(test_photo_as_draw_dir, file)
        #         shutil.copy2(src_file_path, dest_file_path)
        #         if supervised:
        #             label_supervised_test_dir = os.path.join(supervised_test_photo_as_draw_dir, label_dir)
        #             if not os.path.isdir(label_supervised_test_dir):
        #                 os.mkdir(label_supervised_test_dir)
        #             supervised_dest_file_path = os.path.join(label_supervised_test_dir, file)
        #             shutil.copy2(src_file_path, supervised_dest_file_path)
        #         print("{} number {} to test".format(str(label_dir), i))
        #     else:
        #         dest_file_path = os.path.join(os.path.join(train_photo_as_draw_dir, "1"), file)
        #         shutil.copy2(src_file_path, dest_file_path)
        #         if supervised:
        #             label_supervised_train_dir = os.path.join(supervised_train_photo_as_draw_dir, label_dir)
        #             if not os.path.isdir(label_supervised_train_dir):
        #                 os.mkdir(label_supervised_train_dir)
        #             supervised_dest_file_path = os.path.join(label_supervised_train_dir, file)
        #             shutil.copy2(src_file_path, supervised_dest_file_path)
        #         print("{} number {} to train".format(str(label_dir), i))


# create_all_data_drawing("images_final/all_data_drawing", "images_final/labeled_drawing/labels")
# create_all_data_photo("images_final/all_data_photo", "images_final/labeled_photo/labels")

# find_diff_drawing_and_photo( "images_final/labeled_drawing/labels", "images_final/labeled_photo/labels")

# create_drawing_base_dirs("images_final/drawing_base_all", "images_final/labeled_test_train/new_all_data")

# update_photo_dirs("images_final/labeled_drawing", "images_final/labeled_photo")
#
#
update_labels_excel("labels_photo.xlsx", "images_final/labeled_photo/labels")
create_labels_dict("labels_photo.xlsx", "labels_dict_photo.bin")

update_labels_excel("labels_draw.xlsx", "images_final/labeled_drawing/labels")
create_labels_dict("labels_draw.xlsx", "labels_dict_draw.bin")

# update_photo_as_drawing_dir("images_final/labeled_drawing", "images_final/labeled_photo_as_drawing")

test_size = 0.2
for i in range(0, 20):
    create_test_train_experiment("experiment_" + str(i), "images_final/labeled_drawing/", "images_final/labeled_photo/", "images_final/labeled_photo_as_drawing/",
                                 True, test_size)


# not_use_dirs = os.listdir("images_final/drawing_not_use/")
# for dir in not_use_dirs:
#     current_dir = os.path.join("images_final/drawing_not_use", dir)
#     for file in os.listdir(current_dir):
#         os.remove(os.path.join(os.path.join("images_final/labeled_drawing/", dir), file))